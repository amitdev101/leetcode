import csv
from openpyxl import load_workbook

def read_process_csv(file_path):
    """
    Reads a CSV file, prints headers, and processes a specific column.

    :param file_path: Path to the CSV file
    :param column_name: The name of the column to process
    """
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        
        headers = reader.fieldnames
        print(f"Headers: {headers}")
        
        # Process the specified column
        for row in reader:
            print(row['column_name'])
            # Add your processing logic here for the column


def read_process_xlsx(file_path, column_name):
    """
    Reads an XLSX file, prints headers, and processes a specific column by its name.

    :param file_path: Path to the XLSX file
    :param column_name: The name of the column to process
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    
    # Assuming the first row contains headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    print(f"Headers: {headers}")
    
    # Find the index of the column
    column_index = headers.index(column_name) + 1  # Adding 1 because Excel columns start at 1
    
    # Process the specified column
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
        print(row[column_index - 1])  # Subtracting 1 to match Python's 0-based indexing
        # Add your processing logic here for the column
